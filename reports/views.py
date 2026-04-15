from django.http import HttpResponse
from rest_framework.views import APIView
from products.models import Produto
from sales.models import Venda
import openpyxl
from reportlab.pdfgen import canvas
from io import BytesIO


def get_paid_sales_queryset():
    return Venda.objects.filter(status_pagamento='pago').select_related('cliente').order_by('id')

class ExportInventoryExcelView(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estoque Atual"
        
        headers = ['Nome', 'Descrição', 'Estoque Atual', 'Estoque Mínimo', 'Preço de Venda']
        ws.append(headers)
        
        for produto in Produto.objects.all():
            ws.append([produto.sku, produto.descricao, produto.estoque_atual, produto.estoque_minimo, produto.preco_venda])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=estoque.xlsx'
        wb.save(response)
        return response

class ExportSalesReportPDFView(APIView):
    def get(self, request):
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 800, "Relatório de Vendas")
        p.setFont("Helvetica", 10)
        
        headers = ["ID", "Cliente", "Data", "Total"]
        y = 770
        p.drawString(100, y, headers[0])
        p.drawString(150, y, headers[1])
        p.drawString(350, y, headers[2])
        p.drawString(450, y, headers[3])
        p.line(100, y-5, 550, y-5)
        
        y -= 20
        for venda in get_paid_sales_queryset():
            p.drawString(100, y, str(venda.id))
            p.drawString(150, y, str(venda.cliente.nome))
            p.drawString(350, y, venda.data_venda.strftime('%d/%m/%Y'))
            p.drawString(450, y, f"R$ {venda.valor_total}")
            y -= 15
            if y < 50:
                p.showPage()
                y = 800
        
        p.showPage()
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=vendas.pdf'
        return response

class ExportInventoryPDFView(APIView):
    def get(self, request):
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 800, "Relatório de Estoque Atual")
        p.setFont("Helvetica", 10)
        
        headers = ["Nome", "Descrição", "Estoque", "Mínimo", "Preço"]
        y = 770
        p.drawString(50, y, headers[0])
        p.drawString(120, y, headers[1])
        p.drawString(350, y, headers[2])
        p.drawString(420, y, headers[3])
        p.drawString(490, y, headers[4])
        p.line(50, y-5, 550, y-5)
        
        y -= 20
        for produto in Produto.objects.all():
            p.drawString(50, y, str(produto.sku))
            p.drawString(120, y, str(produto.descricao)[:40])
            p.drawString(350, y, str(produto.estoque_atual))
            p.drawString(420, y, str(produto.estoque_minimo))
            p.drawString(490, y, f"R$ {produto.preco_venda}")
            y -= 15
            if y < 50:
                p.showPage()
                y = 800
        
        p.showPage()
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=estoque.pdf'
        return response

class ExportSalesExcelView(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendas"
        headers = ['ID', 'Cliente', 'Data', 'Total']
        ws.append(headers)
        for venda in get_paid_sales_queryset():
            ws.append([venda.id, venda.cliente.nome, venda.data_venda.strftime('%d/%m/%Y'), float(venda.valor_total)])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=vendas.xlsx'
        wb.save(response)
        return response

class ExportFinancialPDFView(APIView):
    def get(self, request):
        from sales.models import ContaReceber
        from purchases.models import ContaPagar
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 800, "Resumo Financeiro")
        p.setFont("Helvetica", 10)
        
        y = 770
        p.drawString(100, y, "CONTAS A RECEBER (PENDENTES)")
        y -= 20
        for cr in ContaReceber.objects.filter(status='pendente'):
            p.drawString(100, y, f"Cliente: {cr.cliente.nome} - Valor: R$ {cr.valor} - Venc: {cr.data_vencimento}")
            y -= 15
            
        y -= 30
        p.drawString(100, y, "CONTAS A PAGAR (PENDENTES)")
        y -= 20
        for cp in ContaPagar.objects.filter(status='pendente'):
            p.drawString(100, y, f"Fornecedor: {cp.fornecedor.razao_social} - Valor: R$ {cp.valor} - Venc: {cp.data_vencimento}")
            y -= 15
            
        p.showPage()
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=financeiro.pdf'
        return response

class ExportFinancialExcelView(APIView):
    def get(self, request):
        from sales.models import ContaReceber
        from purchases.models import ContaPagar
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financeiro"
        ws.append(['Tipo', 'Pessoa', 'Valor', 'Vencimento', 'Status'])
        for cr in ContaReceber.objects.all():
            ws.append(['Receber', cr.cliente.nome, float(cr.valor), cr.data_vencimento, cr.status])
        for cp in ContaPagar.objects.all():
            ws.append(['Pagar', cp.fornecedor.razao_social, float(cp.valor), cp.data_vencimento, cp.status])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=financeiro.xlsx'
        wb.save(response)
        return response
